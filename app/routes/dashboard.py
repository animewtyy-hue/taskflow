from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from app import db
from app.models import Task, Project, Comment, User, Notification
#_____ Helper: إنشاء إشعار________________________
def create_notification(user_id, message, link=None):
    notif = Notification(
        user_id = user_id,
        message = message,
        link = link
    )
    db.session.add(notif)
from datetime import datetime
from flask import make_response
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

dashboard = Blueprint('dashboard', __name__)

@dashboard.route('/dashboard')
@login_required
def index():
    total       = Task.query.filter_by(user_id=current_user.id).count()
    todo        = Task.query.filter_by(user_id=current_user.id, status='todo').count()
    in_progress = Task.query.filter_by(user_id=current_user.id, status='in_progress').count()
    done        = Task.query.filter_by(user_id=current_user.id, status='done').count()
    recent_tasks = Task.query.filter_by(user_id=current_user.id)\
                             .order_by(Task.created_at.desc()).limit(5).all()
    projects = Project.query.all()
    stats = {
        'total': total,
        'todo': todo,
        'in_progress': in_progress,
        'done': done
    }
    return render_template('dashboard.html',
                           stats=stats,
                           recent_tasks=recent_tasks,
                           projects=projects)

@dashboard.route('/tasks')
@login_required
def tasks():
    status   = request.args.get('status', 'all')
    priority = request.args.get('priority', 'all')
    query = Task.query.filter_by(user_id=current_user.id)
    if status   != 'all': query = query.filter_by(status=status)
    if priority != 'all': query = query.filter_by(priority=priority)
    tasks = query.order_by(Task.created_at.desc()).all()
    return render_template('tasks.html', tasks=tasks,
                           status=status, priority=priority)

@dashboard.route('/tasks/new', methods=['GET', 'POST'])
@login_required
def new_task():
    if request.method == 'POST':
        due_date = None
        if request.form.get('due_date'):
            due_date = datetime.strptime(request.form['due_date'], '%Y-%m-%d')

        task = Task(
            title       = request.form['title'],
            description = request.form.get('description', ''),
            status      = request.form.get('status', 'todo'),
            priority    = request.form.get('priority', 'medium'),
            due_date    = due_date,
            user_id     = current_user.id,
            project_id  = request.form.get('project_id') or None
        )
        db.session.add(task)
        db.session.flush()
        create_notification(
            current_user.id,
            f'New task added: {task.title}',
            url_for('dashboard.task_detail', id=task.id)
        )
        db.session.commit()
        flash(' تمت إضافة المهمة بنجاح!', 'success')
        return redirect(url_for('dashboard.tasks'))

    selected_project = request.args.get('project_id')
    projects = Project.query.all()
    return render_template('new_task.html',
                               projects=projects,
                               selected_project=selected_project)

@dashboard.route('/tasks/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_task(id):
    task = Task.query.get_or_404(id)
    if task.user_id != current_user.id and current_user.role != 'admin':
        flash('غير مصرح لك بتعديل هذه المهمة', 'danger')
        return redirect(url_for('dashboard.tasks'))
    if request.method == 'POST':
        task.title       = request.form['title']
        task.description = request.form.get('description', '')
        task.status      = request.form.get('status', 'todo')
        task.priority    = request.form.get('priority', 'medium')
        if request.form.get('due_date'):
            task.due_date = datetime.strptime(request.form['due_date'], '%Y-%m-%d')
        create_notification(
                    current_user.id,
                    f'Task updated: {task.title}',
                    url_for('dashboard.task_detail', id=task.id)
        )
        db.session.commit()
        flash('تم تحديث المهمة!', 'success')
        return redirect(url_for('dashboard.tasks'))
    projects = Project.query.all()
    return render_template('edit_task.html', task=task, projects=projects)

@dashboard.route('/tasks/<int:id>/delete', methods=['POST'])
@login_required
def delete_task(id):
    task = Task.query.get_or_404(id)
    if task.user_id != current_user.id and current_user.role != 'admin':
        flash('غير مصرح لك بحذف هذه المهمة', 'danger')
        return redirect(url_for('dashboard.tasks'))
    db.session.delete(task)
    db.session.commit()
    flash(' تم حذف المهمة', 'warning')
    return redirect(url_for('dashboard.tasks'))
@dashboard.route('/projects')
@login_required
def projects():
    projects = Project.query.all()
    return render_template('projects.html', projects=projects)

@dashboard.route('/projects/new', methods=['GET', 'POST'])
@login_required
def new_project():
    if request.method == 'POST':
        project = Project(
            name        = request.form['name'],
            description = request.form.get('description', '')
        )
        db.session.add(project)
        db.session.commit()
        flash(' تم إنشاء المشروع بنجاح!', 'success')
        return redirect(url_for('dashboard.projects'))
    return render_template('new_project.html')

@dashboard.route('/projects/<int:id>')
@login_required
def project_detail(id):
    project = Project.query.get_or_404(id)
    tasks   = Task.query.filter_by(project_id=id).all()
    total   = len(tasks)
    done    = len([t for t in tasks if t.status == 'done'])
    percent = round((done / total) * 100) if total > 0 else 0
    return render_template('project_detail.html',
                           project=project,
                           tasks=tasks,
                           total=total,
                           done=done,
                           percent=percent)

@dashboard.route('/projects/<int:id>/delete', methods=['POST'])
@login_required
def delete_project(id):
    project = Project.query.get_or_404(id)
    Task.query.filter_by(project_id=id).update({'project_id': None})
    db.session.delete(project)
    db.session.commit()
    flash(' تم حذف المشروع', 'warning')
    return redirect(url_for('dashboard.projects'))
# ─── الرسوم البيانية ───────────────────────────────────

@dashboard.route('/charts')
@login_required
def charts():
    uid = current_user.id

    # إحصائيات الحالة
    todo        = Task.query.filter_by(user_id=uid, status='todo').count()
    in_progress = Task.query.filter_by(user_id=uid, status='in_progress').count()
    done        = Task.query.filter_by(user_id=uid, status='done').count()

    # إحصائيات الأولوية
    high   = Task.query.filter_by(user_id=uid, priority='high').count()
    medium = Task.query.filter_by(user_id=uid, priority='medium').count()
    low    = Task.query.filter_by(user_id=uid, priority='low').count()

    # إحصائيات المشاريع
    projects = Project.query.all()
    project_names  = [p.name for p in projects]
    project_tasks  = [Task.query.filter_by(project_id=p.id).count()
                      for p in projects]

    return render_template('charts.html',
        todo=todo, in_progress=in_progress, done=done,
        high=high, medium=medium, low=low,
        project_names=project_names,
        project_tasks=project_tasks
    )
# ─── التعليقات ────────────────────────────────────────

@dashboard.route('/tasks/<int:id>/detail')
@login_required
def task_detail(id):
    task     = Task.query.get_or_404(id)
    comments = Comment.query.filter_by(task_id=id)\
                            .order_by(Comment.created_at.desc()).all()
    return render_template('task_detail.html',
                           task=task,
                           comments=comments)

@dashboard.route('/tasks/<int:id>/comment', methods=['POST'])
@login_required
def add_comment(id):
    task    = Task.query.get_or_404(id)
    content = request.form.get('content', '').strip()

    if not content:
        flash(' التعليق لا يمكن أن يكون فارغاً', 'warning')
        return redirect(url_for('dashboard.task_detail', id=id))

    comment = Comment(
        content = content,
        user_id = current_user.id,
        task_id = id
    )
    db.session.add(comment)
    db.session.flush()
    create_notification(
        current_user.id,
        f'New comment on: {task.title}',
        url_for('dashboard.task_detail', id=id)
    )
    db.session.commit()
    flash(' تم إضافة التعليق!', 'success')
    return redirect(url_for('dashboard.task_detail', id=id))

@dashboard.route('/comments/<int:id>/delete', methods=['POST'])
@login_required
def delete_comment(id):
    comment = Comment.query.get_or_404(id)
    task_id = comment.task_id

    if comment.user_id != current_user.id and current_user.role != 'admin':
        flash(' غير مصرح لك بحذف هذا التعليق', 'danger')
        return redirect(url_for('dashboard.task_detail', id=task_id))

    db.session.delete(comment)
    db.session.commit()
    flash(' تم حذف التعليق', 'warning')
    return redirect(url_for('dashboard.task_detail', id=task_id))
# ─── الملف الشخصي ─────────────────────────────────────

@dashboard.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email    = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()

        # التحقق من عدم تكرار الاسم
        existing_user = User.query.filter_by(username=username).first()
        if existing_user and existing_user.id != current_user.id:
            flash(' اسم المستخدم مستخدم بالفعل', 'warning')
            return redirect(url_for('dashboard.profile'))

        # التحقق من عدم تكرار البريد
        existing_email = User.query.filter_by(email=email).first()
        if existing_email and existing_email.id != current_user.id:
            flash(' البريد الإلكتروني مستخدم بالفعل', 'warning')
            return redirect(url_for('dashboard.profile'))

        current_user.username = username
        current_user.email    = email

        # تغيير كلمة المرور إذا أدخلها
        if password:
            from werkzeug.security import generate_password_hash
            current_user.password = generate_password_hash(password)
            flash(' تم تحديث البيانات وكلمة المرور!', 'success')
        else:
            flash(' تم تحديث البيانات بنجاح!', 'success')

        db.session.commit()
        return redirect(url_for('dashboard.profile'))

    # إحصائيات المستخدم
    total    = Task.query.filter_by(user_id=current_user.id).count()
    done     = Task.query.filter_by(user_id=current_user.id,
                                    status='done').count()
    comments = Comment.query.filter_by(user_id=current_user.id).count()

    stats = {
        'total':    total,
        'done':     done,
        'comments': comments
    }
    return render_template('profile.html', stats=stats)    
    
# ─── تصدير PDF ────────────────────────────────────────

@dashboard.route('/export/pdf')
@login_required
def export_pdf():
    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story  = []

    # العنوان
    title = Paragraph(
        f"<b>TaskFlow - Tasks Report</b>",
        styles['Title']
    )
    story.append(title)
    story.append(Spacer(1, 20))

    # معلومات المستخدم
    info = Paragraph(
        f"User: {current_user.username} | "
        f"Date: {datetime.now().strftime('%Y-%m-%d')}",
        styles['Normal']
    )
    story.append(info)
    story.append(Spacer(1, 20))

    # إحصائيات
    uid         = current_user.id
    total       = Task.query.filter_by(user_id=uid).count()
    done        = Task.query.filter_by(user_id=uid, status='done').count()
    in_progress = Task.query.filter_by(user_id=uid, status='in_progress').count()
    todo        = Task.query.filter_by(user_id=uid, status='todo').count()

    stats_data = [
        ['Total', 'Done', 'In progress', 'Todo'],
        [str(total), str(done), str(in_progress), str(todo)]
    ]

    stats_table = Table(stats_data, colWidths=[120, 120, 120, 120])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6366f1')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('FONTSIZE',   (0,0), (-1,-1), 12),
        ('ROWBACKGROUNDS', (0,1), (-1,-1),
         [colors.HexColor('#f8fafc'), colors.white]),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0')),
        ('PADDING', (0,0), (-1,-1), 10),
    ]))

    story.append(stats_table)
    story.append(Spacer(1, 30))

    # جدول المهام
    tasks = Task.query.filter_by(user_id=uid)\
                      .order_by(Task.created_at.desc()).all()

    if tasks:
        heading = Paragraph('<b> Tasks List</b>', styles['Heading2'])
        story.append(heading)
        story.append(Spacer(1, 10))

        table_data = [['#', 'Task', 'Priority', 'Status', ' Date']]

        for i, task in enumerate(tasks, 1):
            table_data.append([
                str(i),
                task.title[:40],
                task.priority,
                task.status,
                task.created_at.strftime('%Y-%m-%d')
            ])

        tasks_table = Table(table_data,
                            colWidths=[30, 220, 80, 90, 90])
        tasks_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('FONTSIZE',   (0,0), (-1,-1), 10),
            ('ROWBACKGROUNDS', (0,1), (-1,-1),
             [colors.HexColor('#f8fafc'), colors.white]),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0')),
            ('PADDING', (0,0), (-1,-1), 8),
        ]))

        story.append(tasks_table)

    # بناء الـ PDF
    doc.build(story)
    buffer.seek(0)

    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/pdf'
    response.headers['Content-Disposition'] = \
        f'attachment; filename=taskflow_report_{datetime.now().strftime("%Y%m%d")}.pdf'
    return response    
# ─── تصدير Excel ──────────────────────────────────────

@dashboard.route('/export/excel')
@login_required
def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks Report"

    # ألوان الهيدر
    header_fill = PatternFill(
        start_color="6366F1",
        end_color="6366F1",
        fill_type="solid"
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
        size=12
    )

    # عناوين الأعمدة
    headers = ['#', 'Task', 'Description',
               'Priority', 'Status', 'Due Date', 'Created At']
    ws.append(headers)

    # تنسيق الهيدر
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')

    # عرض الأعمدة
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    # البيانات
    tasks = Task.query.filter_by(user_id=current_user.id)\
                      .order_by(Task.created_at.desc()).all()

    for i, task in enumerate(tasks, 1):
        ws.append([
            i,
            task.title,
            task.description or '',
            task.priority,
            task.status,
            task.due_date.strftime('%Y-%m-%d') if task.due_date else '',
            task.created_at.strftime('%Y-%m-%d')
        ])

        # تلوين صفوف متناوبة
        fill_color = "F8FAFC" if i % 2 == 0 else "FFFFFF"
        row_fill   = PatternFill(
            start_color=fill_color,
            end_color=fill_color,
            fill_type="solid"
        )
        for col in range(1, len(headers) + 1):
            ws.cell(row=i+1, column=col).fill = row_fill

    # ورقة الإحصائيات
    ws2 = wb.create_sheet("Statistics")
    ws2.append(['Metric', 'Value'])

    uid = current_user.id
    stats = [
        ['Total Tasks',    Task.query.filter_by(user_id=uid).count()],
        ['Done',           Task.query.filter_by(user_id=uid, status='done').count()],
        ['In Progress',    Task.query.filter_by(user_id=uid, status='in_progress').count()],
        ['Todo',           Task.query.filter_by(user_id=uid, status='todo').count()],
        ['High Priority',  Task.query.filter_by(user_id=uid, priority='high').count()],
        ['Medium Priority',Task.query.filter_by(user_id=uid, priority='medium').count()],
        ['Low Priority',   Task.query.filter_by(user_id=uid, priority='low').count()],
    ]

    for stat in stats:
        ws2.append(stat)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 10

    # حفظ وإرسال
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = make_response(buffer.read())
    response.headers['Content-Type'] = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = \
        f'attachment; filename=taskflow_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response    
# ─── الإشعارات ────────────────────────────────────────

@dashboard.route('/notifications')
@login_required
def notifications():
    notifs = Notification.query.filter_by(user_id=current_user.id)\
                               .order_by(Notification.created_at.desc()).all()
    # تحديد الكل كمقروء
    for n in notifs:
        n.is_read = True
    db.session.commit()
    return render_template('notifications.html', notifications=notifs)

@dashboard.route('/notifications/read/<int:id>')
@login_required
def read_notification(id):
    notif = Notification.query.get_or_404(id)
    notif.is_read = True
    db.session.commit()
    if notif.link:
        return redirect(notif.link)
    return redirect(url_for('dashboard.notifications'))

@dashboard.route('/notifications/clear', methods=['POST'])
@login_required
def clear_notifications():
    Notification.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()
    flash(' تم مسح كل الإشعارات', 'warning')
    return redirect(url_for('dashboard.notifications'))    
    
# ─── Kanban Board ─────────────────────────────────────

@dashboard.route('/kanban')
@login_required
def kanban():
    todo        = Task.query.filter_by(
                    user_id=current_user.id,
                    status='todo').all()
    in_progress = Task.query.filter_by(
                    user_id=current_user.id,
                    status='in_progress').all()
    done        = Task.query.filter_by(
                    user_id=current_user.id,
                    status='done').all()

    return render_template('kanban.html',
                           todo=todo,
                           in_progress=in_progress,
                           done=done)    